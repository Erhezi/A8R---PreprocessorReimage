"""Phase 1 — pre-check error/warning report.

Lets the user download a per-row summary of every PC1 error and warning
when there are too many issues to address inline. The user fixes the
underlying spreadsheet offline and re-uploads.

One row per active TaskItem (DELETED_PC1 excluded). Header-level errors
(item_id IS NULL) are emitted as a synthetic Input Ref = 0 row so vendor
ID issues etc. still surface in the file.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..common.utils import ny_now
from ..db import task_repo
from ..state import Status

HEADERS = [
    "Mfg Part Num (Input)",
    "Vendor Part Num (Input)",
    "Buyer Part Num (Input)",
    "Description (Input)",
    "Contract Price (Input)",
    "UOM (Input)",
    "QOE (Input)",
    "Effective Date (Input)",
    "Expiration Date (Input)",
    "Contract ID (Input)",
    "ERP Vendor ID (Input)",
    "Organization (Input)",
    "Manufacturer (Input)",
    "Input Ref",
    "With Error (Y/N)",
    "Total Error Count",
    "Errors",
    "Error Notes",
    "With Warnings (Y/N)",
    "Total Warning Count",
    "Warnings",
    "Warning Notes",
]

WRAP_HEADERS = {"Errors", "Error Notes", "Warnings", "Warning Notes", "Description (Input)"}

# PC1 records both errors and warnings into PreCheckError. These types are the
# warning-severity codes emitted by intake_service; everything else is an error.
WARNING_TYPES = {"QOE_UOM_WARNING", "DUPLICATE_MFG_REDUCED", "DUPLICATE_VENDORITEM_REDUCED"}


def _is_warning(error_type: str) -> bool:
    code = (error_type or "").upper()
    if code in WARNING_TYPES:
        return True
    return code.endswith("_WARNING") or code.endswith("_REDUCED")


def _to_date_str(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%m/%d/%Y")
    text = str(value).strip()
    if "T" in text:
        text = text.split("T", 1)[0]
    return text


def _safe_str(value) -> str:
    return "" if value is None else str(value)


def _safe_number(value):
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return text


def _build_filename(task) -> str:
    contract = (getattr(task, "contract_number", None) or "NA").strip() or "NA"
    vendor = (getattr(task, "vendor_id", None) or "NA").strip() or "NA"
    today = ny_now().strftime("%Y%m%d")
    contract = re.sub(r"[^A-Za-z0-9_.-]", "_", contract)
    vendor = re.sub(r"[^A-Za-z0-9_.-]", "_", vendor)
    return f"precheck_report_{contract}_{vendor}_{today}.xlsx"


def _row_from_item(item, recs: list, task_fields: dict) -> dict:
    err_recs = [r for r in recs if not _is_warning(r.error_type)]
    warn_recs = [r for r in recs if _is_warning(r.error_type)]
    price = float(item.unit_price) if item.unit_price is not None else None
    return {
        "Mfg Part Num (Input)": _safe_str(item.mfg_catalog_num),
        "Vendor Part Num (Input)": _safe_str(item.vendor_catalog_num),
        "Buyer Part Num (Input)": "",
        "Description (Input)": _safe_str(item.description),
        "Contract Price (Input)": _safe_number(price),
        "UOM (Input)": _safe_str(item.uom),
        "QOE (Input)": _safe_number(item.qoe),
        "Effective Date (Input)": task_fields["effective_date"],
        "Expiration Date (Input)": task_fields["expiration_date"],
        "Contract ID (Input)": task_fields["contract_id"],
        "ERP Vendor ID (Input)": task_fields["vendor_id"],
        "Organization (Input)": task_fields["organization"],
        "Manufacturer (Input)": task_fields["manufacturer"],
        "Input Ref": _safe_number(item.file_row) if item.file_row else item.item_id,
        "With Error (Y/N)": "Y" if err_recs else "N",
        "Total Error Count": len(err_recs),
        "Errors": "\n".join(r.error_type or "" for r in err_recs),
        "Error Notes": "\n".join(r.error_detail or "" for r in err_recs),
        "With Warnings (Y/N)": "Y" if warn_recs else "N",
        "Total Warning Count": len(warn_recs),
        "Warnings": "\n".join(r.error_type or "" for r in warn_recs),
        "Warning Notes": "\n".join(r.error_detail or "" for r in warn_recs),
    }


def _row_from_header(header_records: list, task_fields: dict) -> dict:
    err_recs = [r for r in header_records if not _is_warning(r.error_type)]
    warn_recs = [r for r in header_records if _is_warning(r.error_type)]
    return {
        "Mfg Part Num (Input)": "",
        "Vendor Part Num (Input)": "",
        "Buyer Part Num (Input)": "",
        "Description (Input)": "(header-level issues)",
        "Contract Price (Input)": "",
        "UOM (Input)": "",
        "QOE (Input)": "",
        "Effective Date (Input)": task_fields["effective_date"],
        "Expiration Date (Input)": task_fields["expiration_date"],
        "Contract ID (Input)": task_fields["contract_id"],
        "ERP Vendor ID (Input)": task_fields["vendor_id"],
        "Organization (Input)": task_fields["organization"],
        "Manufacturer (Input)": task_fields["manufacturer"],
        "Input Ref": 0,
        "With Error (Y/N)": "Y" if err_recs else "N",
        "Total Error Count": len(err_recs),
        "Errors": "\n".join(r.error_type or "" for r in err_recs),
        "Error Notes": "\n".join(r.error_detail or "" for r in err_recs),
        "With Warnings (Y/N)": "Y" if warn_recs else "N",
        "Total Warning Count": len(warn_recs),
        "Warnings": "\n".join(r.error_type or "" for r in warn_recs),
        "Warning Notes": "\n".join(r.error_detail or "" for r in warn_recs),
    }


def build_excel(task_id: str) -> tuple[str, io.BytesIO]:
    task = task_repo.get_task(task_id)
    if not task:
        raise ValueError(f"Task {task_id} not found.")

    items = task_repo.get_items(task_id)
    items = [i for i in items if (i.status or "") != Status.DELETED_PC1]
    errors = task_repo.get_precheck_errors(task_id, phase="PC1")
    # Resolved records have already been addressed (e.g. via manual pass), so
    # they shouldn't show up in a "fix offline" report.
    errors = [e for e in errors if not e.resolved]

    by_item: dict[int, list] = {}
    header_records: list = []
    for e in errors:
        if e.item_id is None:
            header_records.append(e)
        else:
            by_item.setdefault(e.item_id, []).append(e)

    task_fields = {
        "contract_id": _safe_str(getattr(task, "contract_number", "")),
        "vendor_id": _safe_str(getattr(task, "vendor_id", "")),
        "organization": _safe_str(getattr(task, "organization", "")),
        "manufacturer": _safe_str(getattr(task, "oem_name", "")),
        "effective_date": _to_date_str(getattr(task, "contract_start_date", None)),
        "expiration_date": _to_date_str(getattr(task, "contract_end_date", None)),
    }

    rows: list[dict] = []
    if header_records:
        rows.append(_row_from_header(header_records, task_fields))

    items_sorted = sorted(items, key=lambda x: (x.file_row or 10**9, x.item_id))
    for item in items_sorted:
        rows.append(_row_from_item(item, by_item.get(item.item_id, []), task_fields))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "precheck_report"

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, header in enumerate(HEADERS, start=1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=row.get(header, ""))
            if header in WRAP_HEADERS:
                cell.alignment = wrap

    for col_idx, header in enumerate(HEADERS, start=1):
        max_len = len(str(header))
        for row in rows:
            v = row.get(header, "")
            if v is None:
                continue
            longest_line = max((len(line) for line in str(v).split("\n")), default=0)
            if longest_line > max_len:
                max_len = longest_line
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 12), 60
        )

    worksheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return _build_filename(task), buffer
