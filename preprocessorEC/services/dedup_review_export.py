"""Phase 5 — v1.0-compatible "dedup output to review" excel report.

Reads ACCEPTED CCX matches out of ``PreprocessorTaskItemForDecision``
(joined to ``CCXSyncedContractLine`` for matched-side effective /
expiration dates) and produces the multi-sheet review workbook that
sourcing reviewers consumed in v1.0:

  - ``quick_line_count``: per (Organization, ContractID, ERPVendorID)
    counts of total CCX lines vs. matched (overlap) lines.
  - One sheet per matched contract — sheet name = ContractID, with
    ``_1``, ``_2`` suffixes when the same ContractID appears under
    multiple (Organization, ERPVendorID) keys.
"""

from __future__ import annotations

import io
import re
from collections import OrderedDict
from datetime import date, datetime
from typing import Optional

from sqlalchemy import bindparam
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..common.utils import ny_now
from ..db import task_repo
from ..db.engine import get_sqlserver_engine
from ..db.sql_loader import load_query
from . import dedup_review_rules

MATCHED_HEADERS = [
    "Mfg Part Number",
    "Vendor Part Number",
    "Buyer Part Num",
    "Description",
    "Contract Price",
    "UOM",
    "QOE",
    "Effective Date",
    "Expiration Date",
    "Contract ID",
    "ERP Vendor ID",
    "Organization",
    "Action",
    "Notes",
]
INPUT_HEADERS = [
    "Mfg Part Num (Input)",
    "Vendor Part Num (Input)",
    "Description (Input)",
    "Contract Price (Input)",
    "UOM (Input)",
    "QOE (Input)",
    "Contract ID (Input)",
    "ERP Vendor ID (Input)",
    "Organization (Input)",
    "Infor Item #",
]
ALL_HEADERS = MATCHED_HEADERS + INPUT_HEADERS

# view_by_input sheet: input columns first (light blue), matched second
# (yellow). No Action/Notes on this sheet — it's a flat input-centric view.
VIEW_BY_INPUT_INPUT_HEADERS = [
    "Mfg Part Num (Input)",
    "Vendor Part Num (Input)",
    "Buyer Part Num (Input)",
    "Description (Input)",
    "UOM (Input)",
    "QOE (Input)",
    "Effective Date (Input)",
    "Expiration Date (Input)",
    "Contract ID (Input)",
    "ERP Vendor ID (Input)",
    "Organization (Input)",
    "Infor Item #",
    "Infor Item BuyUOM Options",
    "Valid BuyUOM (Y/N)",
    "Input Ref",
    "Dup Matched (Y/N)",
    "Total Matched Lines",
]
VIEW_BY_INPUT_MATCHED_HEADERS = [
    "Mfg Part Number",
    "Vendor Part Number",
    "Buyer Part Num",
    "Description",
    "Contract Price",
    "UOM",
    "QOE",
    "Effective Date",
    "Expiration Date",
    "Contract ID",
    "ERP Vendor ID",
    "Organization",
]
VIEW_BY_INPUT_HEADERS = VIEW_BY_INPUT_INPUT_HEADERS + VIEW_BY_INPUT_MATCHED_HEADERS
VIEW_BY_INPUT_SHEET_NAME = "view_by_input"

DATE_COLUMNS = {"Effective Date", "Expiration Date"}

YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
INVALID_BUY_UOM_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HEADER_FONT = Font(bold=True)
HYPERLINK_FONT = Font(color="0563C1", underline="single")


def _normalized_buy_uom_option(uom, qoe) -> Optional[str]:
    """Build the "UOM*QOE" key used to test membership in the Infor buy-UOM
    options string. Mirrors preprocess_service._build_buy_uom_option so
    validation here matches what Phase 3 wrote to PreprocessorItemMatching."""
    text = str(uom or "").strip().upper()
    if not text:
        return None
    try:
        conv = int(qoe)
    except (TypeError, ValueError):
        return None
    if conv <= 0:
        return None
    return f"{text}*{conv}"


def _parse_buy_uom_options(value) -> set[str]:
    if not value:
        return set()
    return {chunk.strip().upper() for chunk in str(value).split(",") if chunk.strip()}


def _internal_sheet_link(sheet_name: str) -> str:
    """Build an Excel internal hyperlink target like ``#'Sheet Name'!A1``.

    Single quotes inside the sheet name are doubled per Excel's quoting rules.
    """
    safe = (sheet_name or "").replace("'", "''")
    return f"#'{safe}'!A1"


def _session() -> Session:
    return Session(get_sqlserver_engine())


def _to_date_str(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%m/%d/%Y")
    text = str(value).strip()
    # Trim ISO-like timestamps so the cell shows just the date portion.
    if "T" in text:
        text = text.split("T", 1)[0]
    if " " in text and ":" in text:
        text = text.split(" ", 1)[0]
    try:
        parsed = datetime.fromisoformat(text)
        return parsed.strftime("%m/%d/%Y")
    except ValueError:
        return text


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value)


def _safe_number(value):
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return ""
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return text


def _normalize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no /\\?*[]: characters."""
    cleaned = re.sub(r"[\\/\?\*\[\]:]", "_", name or "")
    return cleaned[:31] or "contract"


def _matched_row_dict(row) -> dict:
    group = row.get("resolution_grouping")
    intention = row.get("task_intention")
    action = dedup_review_rules.action_for(group, intention)
    notes = dedup_review_rules.notes_for(
        group,
        intention,
        matched_contract_id=row.get("contract_id_matched"),
        matched_org_eid=row.get("organization_eid_matched"),
        input_org_eid=row.get("organization_eid_input"),
        matched_source_type=row.get("matched_contract_source_type"),
        input_source_type=row.get("input_contract_source_type"),
        ea_matched=row.get("ea_price_matched"),
        ea_input=row.get("ea_price_input"),
    )
    return {
        "Mfg Part Number": _safe_str(row["manufacturer_number_matched"]),
        "Vendor Part Number": _safe_str(row["vendor_item_matched"]),
        "Buyer Part Num": "",
        "Description": _safe_str(row["item_desc_matched"]),
        "Contract Price": _safe_number(row["contract_price_matched"]),
        "UOM": _safe_str(row["uom_matched"]),
        "QOE": _safe_number(row["qoe_matched"]),
        "Effective Date": _to_date_str(row["effective_date_matched"]),
        "Expiration Date": _to_date_str(row["expiration_date_matched"]),
        "Contract ID": _safe_str(row["contract_id_matched"]),
        "ERP Vendor ID": _safe_str(row["erp_vendor_id_matched"]),
        "Organization": _safe_str(row["organization_matched"]),
        "Action": action,
        "Notes": notes,
        "Mfg Part Num (Input)": _safe_str(row["manufacturer_number_input"]),
        "Vendor Part Num (Input)": _safe_str(row["vendor_item_input"]),
        "Description (Input)": _safe_str(row["item_description_input"]),
        "Contract Price (Input)": _safe_number(row["contract_price_input"]),
        "UOM (Input)": _safe_str(row["uom_input"]),
        "QOE (Input)": _safe_number(row["qoe_input"]),
        "Contract ID (Input)": _safe_str(row["contract_id_input"]),
        "ERP Vendor ID (Input)": _safe_str(row["erp_vendor_id_input"]),
        "Organization (Input)": _safe_str(row["organization_input"]),
        "Infor Item #": _safe_str(row["infor_item_number"]),
    }


def _augment_groups_with_replacements(
    groups: "OrderedDict[str, dict]",
    unmatched_rows: list[dict],
) -> "OrderedDict[str, dict]":
    """Attach unmatched-CCX rows to per-contract sheet groups for REPLACE
    scopes, creating new groups for REPLACE contracts that have no ACCEPTED
    matches at all.

    Rows are tagged ``__replacement__: True`` so the formatting step routes
    them through ``_replacement_unmatched_row_dict`` instead of the regular
    matched-row path (which expects input-side columns that aren't present
    on these CCX-only rows).
    """
    if not unmatched_rows:
        return groups

    by_key: "OrderedDict[tuple, list[dict]]" = OrderedDict()
    for row in unmatched_rows:
        key = (
            row.get("organization_eid_matched") or "",
            row.get("contract_id_matched") or "",
            row.get("erp_vendor_id_matched") or "",
        )
        tagged = dict(row)
        tagged["__replacement__"] = True
        by_key.setdefault(key, []).append(tagged)

    sheet_name_by_key = {
        (
            payload["key"]["organization_eid"],
            payload["key"]["contract_id"],
            payload["key"]["erp_vendor_id"],
        ): sheet_name
        for sheet_name, payload in groups.items()
    }

    contract_total: dict[str, int] = {}
    for sheet_name, payload in groups.items():
        contract_total[payload["key"]["contract_id"]] = (
            contract_total.get(payload["key"]["contract_id"], 0) + 1
        )
    for key in by_key:
        if key in sheet_name_by_key:
            continue
        contract_total[key[1]] = contract_total.get(key[1], 0) + 1

    contract_counter: dict[str, int] = {}
    used_names = set(groups.keys())

    for key, repl_rows in by_key.items():
        org_eid, contract_id, erp_vendor = key
        existing_sheet = sheet_name_by_key.get(key)
        if existing_sheet:
            groups[existing_sheet]["rows"].extend(repl_rows)
            continue

        # No ACCEPTED matches under this REPLACE scope — create a fresh sheet
        # carrying only unmatched rows. Reuse the same naming convention as
        # _build_sheet_groups so a contract appearing under multiple scopes
        # gets _1/_2 suffixes.
        base_name = contract_id or "no_contract"
        if contract_total.get(contract_id, 0) > 1:
            contract_counter[contract_id] = contract_counter.get(contract_id, 0) + 1
            base_name = f"{base_name}_{contract_counter[contract_id]}"
        sheet_name = _normalize_sheet_name(base_name)
        candidate = sheet_name
        suffix = 1
        while candidate in used_names:
            suffix += 1
            candidate = _normalize_sheet_name(f"{sheet_name}_{suffix}")
        used_names.add(candidate)

        first = repl_rows[0]
        groups[candidate] = {
            "key": {
                "organization_eid": org_eid,
                "contract_id": contract_id,
                "erp_vendor_id": erp_vendor,
                "organization": first.get("organization_matched") or "",
                "contract_id_input": "",
                "erp_vendor_id_input": "",
                "organization_input": "",
            },
            "rows": list(repl_rows),
        }
    return groups


def _build_sheet_groups(rows: list[dict]) -> "OrderedDict[str, dict]":
    """Partition rows by (Organization, ContractID, ERPVendorID) and
    pick a unique sheet name per group.

    Same ContractID under different (Organization, ERPVendorID) gets
    ``_1``, ``_2`` suffixes per the spec.
    """
    groups: "OrderedDict[tuple, list[dict]]" = OrderedDict()
    for row in rows:
        key = (
            row.get("organization_eid_matched") or "",
            row.get("contract_id_matched") or "",
            row.get("erp_vendor_id_matched") or "",
        )
        groups.setdefault(key, []).append(row)

    result: "OrderedDict[str, dict]" = OrderedDict()
    contract_counter: dict[str, int] = {}
    contract_total: dict[str, int] = {}
    for key in groups:
        contract_total[key[1]] = contract_total.get(key[1], 0) + 1

    for key, group_rows in groups.items():
        org_eid, contract_id, erp_vendor = key
        base_name = contract_id or "no_contract"
        if contract_total.get(contract_id, 0) > 1:
            contract_counter[contract_id] = contract_counter.get(contract_id, 0) + 1
            base_name = f"{base_name}_{contract_counter[contract_id]}"
        sheet_name = _normalize_sheet_name(base_name)

        # Disambiguate if normalization collapsed two different names.
        candidate = sheet_name
        suffix = 1
        while candidate in result:
            suffix += 1
            candidate = _normalize_sheet_name(f"{sheet_name}_{suffix}")
        first = group_rows[0] if group_rows else {}
        result[candidate] = {
            "key": {
                "organization_eid": org_eid,
                "contract_id": contract_id,
                "erp_vendor_id": erp_vendor,
                "organization": first.get("organization_matched") or "",
                "contract_id_input": first.get("contract_id_input") or "",
                "erp_vendor_id_input": first.get("erp_vendor_id_input") or "",
                "organization_input": first.get("organization_input") or "",
            },
            "rows": group_rows,
        }
    return result


def _fetch_review_rows(session: Session, task_id: str) -> list[dict]:
    stmt = load_query("export", "dedup_review", query="dedup_review_rows")
    return [dict(r) for r in session.execute(stmt, {"task_id": task_id}).mappings().all()]


def _fetch_view_by_input_rows(session: Session, task_id: str) -> list[dict]:
    stmt = load_query("export", "dedup_review", query="view_by_input_rows")
    return [dict(r) for r in session.execute(stmt, {"task_id": task_id}).mappings().all()]


def _fetch_replacement_unmatched_rows(session: Session, task_id: str) -> list[dict]:
    """CCX lines on REPLACE-marked contracts that did NOT match any input."""
    stmt = load_query("export", "dedup_review", query="replacement_unmatched_lines")
    return [dict(r) for r in session.execute(stmt, {"task_id": task_id}).mappings().all()]


REPLACEMENT_UNMATCHED_NOTES = (
    "check if the item is discontinued, or evaluate if we need put this to "
    "the new contract"
)


def _replacement_unmatched_row_dict(row) -> dict:
    """Format a CCX-only "unmatched on to-be-replaced contract" row.

    Matched columns come from CCX; input columns are blank because there is
    no input line for these items. Action / Notes carry the replacement
    explanation called out in the report spec.
    """
    matched_contract_id = _safe_str(row.get("contract_id_matched"))
    return {
        "Mfg Part Number": _safe_str(row.get("manufacturer_number_matched")),
        "Vendor Part Number": _safe_str(row.get("vendor_item_matched")),
        "Buyer Part Num": "",
        "Description": _safe_str(row.get("item_desc_matched")),
        "Contract Price": _safe_number(row.get("contract_price_matched")),
        "UOM": _safe_str(row.get("uom_matched")),
        "QOE": _safe_number(row.get("qoe_matched")),
        "Effective Date": _to_date_str(row.get("effective_date_matched")),
        "Expiration Date": _to_date_str(row.get("expiration_date_matched")),
        "Contract ID": matched_contract_id,
        "ERP Vendor ID": _safe_str(row.get("erp_vendor_id_matched")),
        "Organization": _safe_str(row.get("organization_matched")),
        "Action": f"Only seen on to-be replaced contract {matched_contract_id}",
        "Notes": REPLACEMENT_UNMATCHED_NOTES,
        "Mfg Part Num (Input)": "",
        "Vendor Part Num (Input)": "",
        "Description (Input)": "",
        "Contract Price (Input)": "",
        "UOM (Input)": "",
        "QOE (Input)": "",
        "Contract ID (Input)": "",
        "ERP Vendor ID (Input)": "",
        "Organization (Input)": "",
        "Infor Item #": "",
    }


def _view_by_input_row_dict(row) -> dict:
    has_match = row.get("dedup_id") is not None
    total = int(row.get("total_matched_lines") or 0)
    options_raw = row.get("infor_buy_uom_options")
    options_set = _parse_buy_uom_options(options_raw)
    expected_option = _normalized_buy_uom_option(
        row.get("uom_to_match_infor_input"), row.get("qoe_input")
    )
    if not options_set or not expected_option:
        valid_buy_uom = ""
    else:
        valid_buy_uom = "Yes" if expected_option in options_set else "No"
    return {
        "Mfg Part Num (Input)": _safe_str(row["manufacturer_number_input"]),
        "Vendor Part Num (Input)": _safe_str(row["vendor_item_input"]),
        "Buyer Part Num (Input)": "",
        "Description (Input)": _safe_str(row["item_description_input"]),
        "UOM (Input)": _safe_str(row["uom_input"]),
        "QOE (Input)": _safe_number(row["qoe_input"]),
        "Effective Date (Input)": _to_date_str(row.get("effective_date_input")),
        "Expiration Date (Input)": _to_date_str(row.get("expiration_date_input")),
        "Contract ID (Input)": _safe_str(row["contract_id_input"]),
        "ERP Vendor ID (Input)": _safe_str(row["erp_vendor_id_input"]),
        "Organization (Input)": _safe_str(row["organization_input"]),
        "Infor Item #": _safe_str(row["infor_item_number"]),
        "Infor Item BuyUOM Options": _safe_str(options_raw),
        "Valid BuyUOM (Y/N)": valid_buy_uom,
        "Input Ref": _safe_number(row["file_row"]),
        "Dup Matched (Y/N)": "Yes" if has_match else "No",
        "Total Matched Lines": total,
        "Mfg Part Number": _safe_str(row["manufacturer_number_matched"]) if has_match else "",
        "Vendor Part Number": _safe_str(row["vendor_item_matched"]) if has_match else "",
        "Buyer Part Num": "",
        "Description": _safe_str(row["item_desc_matched"]) if has_match else "",
        "Contract Price": _safe_number(row["contract_price_matched"]) if has_match else "",
        "UOM": _safe_str(row["uom_matched"]) if has_match else "",
        "QOE": _safe_number(row["qoe_matched"]) if has_match else "",
        "Effective Date": _to_date_str(row["effective_date_matched"]) if has_match else "",
        "Expiration Date": _to_date_str(row["expiration_date_matched"]) if has_match else "",
        "Contract ID": _safe_str(row["contract_id_matched"]) if has_match else "",
        "ERP Vendor ID": _safe_str(row["erp_vendor_id_matched"]) if has_match else "",
        "Organization": _safe_str(row["organization_matched"]) if has_match else "",
    }


def _fetch_contract_totals(
    session: Session, contract_ids: list[str]
) -> dict[tuple, int]:
    contract_ids = sorted({c for c in contract_ids if c})
    if not contract_ids:
        return {}
    stmt = load_query("export", "dedup_review", query="contract_line_counts")
    bound = stmt.bindparams(bindparam("contract_ids", expanding=True))
    out: dict[tuple, int] = {}
    for row in session.execute(bound, {"contract_ids": contract_ids}).mappings().all():
        key = (
            row.get("OrganizationEID") or "",
            row.get("ContractID") or "",
            row.get("ERPVendorID") or "",
        )
        out[key] = int(row.get("LineCnt_CCX") or 0)
    return out


def get_review_data(task_id: str) -> dict:
    """Return the structured payload used to render preview + build excel.

    Shape::

        {
            "task_id": str,
            "filename": str,
            "summary": [
                {"contract_id", "erp_vendor_id", "organization",
                 "organization_eid", "total_lines", "matched_lines",
                 "sheet_name"}, ...
            ],
            "sheets": [
                {"sheet_name", "key": {...}, "rows": [{header: value, ...}], "row_count"},
                ...
            ],
        }
    """
    task = task_repo.get_task(task_id)
    if not task:
        raise ValueError(f"Task {task_id} not found.")

    with _session() as session:
        rows = _fetch_review_rows(session, task_id)
        groups = _build_sheet_groups(rows)
        unmatched_repl_raw = _fetch_replacement_unmatched_rows(session, task_id)
        groups = _augment_groups_with_replacements(groups, unmatched_repl_raw)
        totals = _fetch_contract_totals(
            session, [g["key"]["contract_id"] for g in groups.values()]
        )
        view_by_input_raw = _fetch_view_by_input_rows(session, task_id)

    view_by_input_rows = [_view_by_input_row_dict(r) for r in view_by_input_raw]

    # quick_line_count overview row for the view_by_input sheet:
    # Total Lines = distinct input lines on the task, Matched Lines = distinct
    # inputs with at least one ACCEPTED CCX match. ID columns intentionally
    # blank — this is a task-wide row, not a per-contract one.
    seen_inputs: set = set()
    matched_inputs: set = set()
    for r in view_by_input_raw:
        iid = r.get("input_item_id")
        if iid is None:
            continue
        seen_inputs.add(iid)
        if (r.get("total_matched_lines") or 0) > 0:
            matched_inputs.add(iid)

    summary: list[dict] = [{
        "sheet_name": VIEW_BY_INPUT_SHEET_NAME,
        "contract_id": "",
        "erp_vendor_id": "",
        "organization": "",
        "organization_eid": "",
        "contract_id_input": "",
        "erp_vendor_id_input": "",
        "organization_input": "",
        "total_lines": len(seen_inputs),
        "matched_lines": len(matched_inputs),
    }]
    sheets: list[dict] = []
    for sheet_name, payload in groups.items():
        key = payload["key"]
        formatted_rows: list[dict] = []
        matched_lines = 0
        for raw in payload["rows"]:
            if raw.get("__replacement__"):
                formatted_rows.append(_replacement_unmatched_row_dict(raw))
            else:
                formatted_rows.append(_matched_row_dict(raw))
                matched_lines += 1
        total_lines = totals.get(
            (key["organization_eid"], key["contract_id"], key["erp_vendor_id"]),
            0,
        )
        summary.append({
            "sheet_name": sheet_name,
            "contract_id": key["contract_id"],
            "erp_vendor_id": key["erp_vendor_id"],
            "organization": key["organization"],
            "organization_eid": key["organization_eid"],
            "contract_id_input": key["contract_id_input"],
            "erp_vendor_id_input": key["erp_vendor_id_input"],
            "organization_input": key["organization_input"],
            "total_lines": total_lines,
            "matched_lines": matched_lines,
        })
        sheets.append({
            "sheet_name": sheet_name,
            "key": key,
            "rows": formatted_rows,
            "row_count": len(formatted_rows),
        })

    filename = _build_filename(task)

    return {
        "task_id": task_id,
        "filename": filename,
        "headers": ALL_HEADERS,
        "matched_headers": MATCHED_HEADERS,
        "input_headers": INPUT_HEADERS,
        "summary": summary,
        "sheets": sheets,
        "view_by_input": {
            "sheet_name": VIEW_BY_INPUT_SHEET_NAME,
            "headers": VIEW_BY_INPUT_HEADERS,
            "input_headers": VIEW_BY_INPUT_INPUT_HEADERS,
            "matched_headers": VIEW_BY_INPUT_MATCHED_HEADERS,
            "rows": view_by_input_rows,
            "row_count": len(view_by_input_rows),
        },
    }


def _build_filename(task) -> str:
    contract = (getattr(task, "contract_number", None) or "NA").strip() or "NA"
    vendor = (getattr(task, "vendor_id", None) or "NA").strip() or "NA"
    today = ny_now().strftime("%Y%m%d")
    contract = re.sub(r"[^A-Za-z0-9_.-]", "_", contract)
    vendor = re.sub(r"[^A-Za-z0-9_.-]", "_", vendor)
    return f"dedup_output_to_review_{contract}_{vendor}_{today}.xlsx"


def _write_header_row(worksheet, headers: list[str], fills: list[PatternFill]) -> None:
    for col_idx, (header, fill) in enumerate(zip(headers, fills), start=1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill


def _autofit(worksheet, headers: list[str], rows: list[dict]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in rows:
            v = r.get(header, "")
            if v is None:
                continue
            text = str(v)
            if len(text) > max_len:
                max_len = len(text)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(
            max(max_len + 2, 10), 50
        )


def build_excel(task_id: str) -> tuple[str, io.BytesIO]:
    """Build the dedup-review workbook in-memory. Returns (filename, BytesIO)."""
    data = get_review_data(task_id)

    workbook = Workbook()
    # Replace the default sheet with our quick_line_count sheet.
    summary_ws = workbook.active
    summary_ws.title = "quick_line_count"

    summary_headers = [
        "Contract ID",
        "ERP Vendor ID",
        "Organization",
        "Total Lines",
        "Matched Lines",
        "Contract ID (Input)",
        "ERP Vendor ID (Input)",
        "Organization (Input)",
        "Sheet",
    ]
    summary_fills = (
        [YELLOW_FILL] * 5
        + [LIGHT_BLUE_FILL] * 3
        + [YELLOW_FILL]
    )
    for col_idx, (header, fill) in enumerate(zip(summary_headers, summary_fills), start=1):
        cell = summary_ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = fill
    for r_idx, row in enumerate(data["summary"], start=2):
        summary_ws.cell(row=r_idx, column=1, value=row["contract_id"])
        summary_ws.cell(row=r_idx, column=2, value=row["erp_vendor_id"])
        summary_ws.cell(row=r_idx, column=3, value=row["organization"])
        summary_ws.cell(row=r_idx, column=4, value=row["total_lines"])
        summary_ws.cell(row=r_idx, column=5, value=row["matched_lines"])
        summary_ws.cell(row=r_idx, column=6, value=row["contract_id_input"])
        summary_ws.cell(row=r_idx, column=7, value=row["erp_vendor_id_input"])
        summary_ws.cell(row=r_idx, column=8, value=row["organization_input"])
        sheet_cell = summary_ws.cell(row=r_idx, column=9, value=row["sheet_name"])
        if row.get("sheet_name"):
            sheet_cell.hyperlink = _internal_sheet_link(row["sheet_name"])
            sheet_cell.font = HYPERLINK_FONT
    _autofit(
        summary_ws,
        summary_headers,
        [
            {
                "Contract ID": s["contract_id"],
                "ERP Vendor ID": s["erp_vendor_id"],
                "Organization": s["organization"],
                "Total Lines": s["total_lines"],
                "Matched Lines": s["matched_lines"],
                "Contract ID (Input)": s["contract_id_input"],
                "ERP Vendor ID (Input)": s["erp_vendor_id_input"],
                "Organization (Input)": s["organization_input"],
                "Sheet": s["sheet_name"],
            }
            for s in data["summary"]
        ],
    )

    view_by_input = data.get("view_by_input") or {}
    vbi_rows = view_by_input.get("rows", [])
    vbi_input_headers = view_by_input.get("input_headers", VIEW_BY_INPUT_INPUT_HEADERS)
    vbi_matched_headers = view_by_input.get("matched_headers", VIEW_BY_INPUT_MATCHED_HEADERS)
    vbi_headers = view_by_input.get("headers", VIEW_BY_INPUT_HEADERS)
    vbi_fills = (
        [LIGHT_BLUE_FILL] * len(vbi_input_headers)
        + [YELLOW_FILL] * len(vbi_matched_headers)
    )
    vbi_ws = workbook.create_sheet(
        title=view_by_input.get("sheet_name", VIEW_BY_INPUT_SHEET_NAME)
    )
    _write_header_row(vbi_ws, vbi_headers, vbi_fills)
    for r_idx, row in enumerate(vbi_rows, start=2):
        invalid_buy_uom = row.get("Valid BuyUOM (Y/N)") == "No"
        for c_idx, header in enumerate(vbi_headers, start=1):
            cell = vbi_ws.cell(row=r_idx, column=c_idx, value=row.get(header, ""))
            if invalid_buy_uom:
                cell.fill = INVALID_BUY_UOM_FILL
    _autofit(vbi_ws, vbi_headers, vbi_rows)
    vbi_ws.freeze_panes = "A2"

    fills = [YELLOW_FILL] * len(MATCHED_HEADERS) + [LIGHT_BLUE_FILL] * len(INPUT_HEADERS)

    for sheet in data["sheets"]:
        ws = workbook.create_sheet(title=sheet["sheet_name"])
        _write_header_row(ws, ALL_HEADERS, fills)
        for r_idx, row in enumerate(sheet["rows"], start=2):
            for c_idx, header in enumerate(ALL_HEADERS, start=1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(header, ""))
        _autofit(ws, ALL_HEADERS, sheet["rows"])
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return data["filename"], buffer
